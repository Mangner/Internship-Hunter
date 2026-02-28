import asyncio
import io
import discord
from discord.ext import commands, tasks
from selenium import webdriver
from openpyxl import Workbook
from openpyxl.styles import Font
from database import Database
from Pages.career_page import CareerPage
from Parsers.offerParser import OfferParser
from Repository.offerRepository import OfferRepository
from Repository.subscriptionRepository import SubscriptionRepository
from Services.offerService import OfferService


class OfferCog(commands.Cog):
    def __init__(self, bot: commands.Bot, db_url: str, start_loop: bool = True):
        self._bot = bot
        self._db_url = db_url
        if start_loop:
            self._poll_new_offers.start()

    def cog_unload(self):
        self._poll_new_offers.cancel()

    @commands.hybrid_command(name="register")
    async def register(self, ctx: commands.Context):
        db = Database(self._db_url)
        db.create_tables()
        session = db.get_session()
        try:
            repository = SubscriptionRepository(session)
            repository.create_if_missing(channel_id=ctx.channel.id, guild_id=ctx.guild.id if ctx.guild else None)
        finally:
            session.close()
            db.dispose()

        await ctx.send("Ten kanal zostal zarejestrowany do powiadomien o nowych ofertach.")

    @commands.hybrid_command(name="unregister")
    async def unregister(self, ctx: commands.Context):
        db = Database(self._db_url)
        db.create_tables()
        session = db.get_session()
        try:
            repository = SubscriptionRepository(session)
            removed = repository.delete_by_channel_id(channel_id=ctx.channel.id)
        finally:
            session.close()
            db.dispose()

        if removed:
            await ctx.send("Ten kanal zostal wypisany z powiadomien.")
        else:
            await ctx.send("Ten kanal nie byl zarejestrowany.")

    @commands.hybrid_command(name="showoffers", aliases=["showOffers"])
    async def show_offers(self, ctx: commands.Context):
        db = Database(self._db_url)
        db.create_tables()
        session = db.get_session()
        try:
            repository = OfferRepository(session)
            offers = repository.get_all()
        finally:
            session.close()
            db.dispose()

        if not offers:
            await ctx.send("Brak ofert w bazie.")
            return

        dm = ctx.author
        try:
            await dm.send("Wysylam liste ofert w pliku XLSX.")
            buffer = io.BytesIO()
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Offers"
            sheet.append(["name", "add_date", "exp_date", "href"])
            for offer in offers:
                sheet.append([offer.offer_name, offer.add_date, offer.exp_date, offer.href])

            for row in range(2, sheet.max_row + 1):
                cell = sheet.cell(row=row, column=4)
                if cell.value:
                    cell.hyperlink = cell.value
                    cell.font = Font(color="0000FF", underline="single")

            workbook.save(buffer)
            buffer.seek(0)
            file = discord.File(fp=buffer, filename="offers.xlsx")
            await dm.send(file=file)
        except discord.Forbidden:
            await ctx.send("Nie moge wyslac DM. Sprawdz ustawienia prywatnosci.")

    @commands.hybrid_command(name="help")
    async def help(self, ctx: commands.Context):
        await ctx.send(
            "Dostepne komendy:\n"
            "/register - rejestruje kanal do powiadomien\n"
            "/unregister - wypisuje kanal z powiadomien\n"
            "/showoffers - wysyla liste ofert na DM"
        )

    @tasks.loop(minutes=10)
    async def _poll_new_offers(self):
        await self._bot.wait_until_ready()
        saved_offers = await asyncio.to_thread(self._scrape_and_save_new_offers)

        if not saved_offers:
            return

        await self._notify_subscribers(saved_offers)

    async def _notify_subscribers(self, offers, delay_seconds: int = 5):
        db = Database(self._db_url)
        db.create_tables()
        session = db.get_session()
        try:
            subscription_repo = SubscriptionRepository(session)
            subscriptions = subscription_repo.get_all()
        finally:
            session.close()
            db.dispose()

        print(f"[notify] Found {len(subscriptions)} subscription(s), {len(offers)} offer(s) to send.")

        for subscription in subscriptions:
            print(f"[notify] Looking up channel {subscription.channel_id}...")
            channel = self._bot.get_channel(subscription.channel_id)
            if channel is None:
                print(f"[notify] Channel not in cache, fetching from API...")
                try:
                    channel = await self._bot.fetch_channel(subscription.channel_id)
                except discord.NotFound as e:
                    print(f"[notify] Channel {subscription.channel_id} not found: {e.status} {e.text}")
                    continue
                except discord.Forbidden as e:
                    print(f"[notify] No access to channel {subscription.channel_id}: {e.status} {e.text}")
                    continue
                except discord.HTTPException as e:
                    print(f"[notify] HTTP error for channel {subscription.channel_id}: {e.status} {e.text}")
                    continue

            print(f"[notify] Sending to channel: {channel} ({channel.id})")
            for offer in offers:
                await channel.send(f"Nowa oferta: {offer.offer_name}\n{offer.href}")
                print(f"[notify] Sent: {offer.offer_name}")
                if delay_seconds:
                    await asyncio.sleep(delay_seconds)

    def _scrape_and_save_new_offers(self):
        db = Database(self._db_url)
        db.create_tables()
        session = db.get_session()
        driver = webdriver.Chrome()
        try:
            repository = OfferRepository(session)
            parser = OfferParser()
            service = OfferService(repository, parser)
            page = CareerPage(driver)
            page.open()
            raw_offers = page.get_new_offers(repository.exists_by_href)
            saved = service.save_new_offers(raw_offers)
            for offer in saved:
                session.refresh(offer)
                session.expunge(offer)
            return saved
        finally:
            session.close()
            db.dispose()
            driver.quit()


async def setup(bot: commands.Bot):
    await bot.add_cog(OfferCog(bot, db_url=bot.db_url))
